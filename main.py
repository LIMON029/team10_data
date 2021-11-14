#-*-coding:utf-8-*-
import requests
import numpy as np
from openpyxl import Workbook, load_workbook

# 시도코드와 이에 따른 시도명을 구하는 함수 - (시도코드 : 시도명) 형태
def getSidoNames():
    f = open('sidoCode.txt', 'r', encoding='utf-8')
    lines = f.readlines()

    sidoNames = {}

    for line in lines:
        sido = line.split()
        sidoNames[sido[0]] = sido[1]

    f.close()
    return sidoNames

def getSgguNames():
    f = open('sgguCode.txt', 'r', encoding='utf-8')
    lines = f.readlines()

    sgguNames = {}

    for line in lines:
        codes = line.split()
        sgguNames[codes[0]] = codes[1]

    f.close()
    return sgguNames

# 약효분류를 구하는 함수 - (100 : 신경계 감각기관용 의약품)형태
def getMeftDiv():
    f = open('meftDivNos.txt', 'r', encoding='utf-8')
    lines = f.readlines()

    meftDiv = {}

    for line in lines:
        codes = line.split()
        # 약효분류코드가 저장된 형태 : 100 신경계 감각기관용 의약품
        meftDiv[codes[0]] = codes[1]

    f.close()
    return meftDiv

# 약효분류코드를 구하는 함수(실제 이름을 제외한 코드번호만)
def getMeftDivNos():
    meftDiv = getMeftDiv()
    meftDivNos = []

    for meftDivNo in meftDiv.keys():
        meftDivNos.append(meftDivNo)

    return meftDivNos

# 연월을 구하는 함수(예: 201801)
def getDiagYms():
    diagYms = list(map(str, list(range(201712, 201700, -1))))
    return diagYms

# 추출한 xml문자열에서 필요한 내용을 추출할 부분
# totUseQty : 사용량
def parse(xmlData):
    index = -1
    count = 0
    price = 0
    while True:
        index = xmlData.find('<msupUseAmt>', index + 1)
        if index == -1:
            break
        price_start = index
        price_end = xmlData.find('</msupUseAmt>', index)
        count_start = xmlData.find('<totUseQty>', index)
        count_end = xmlData.find('</totUseQty>', index)
        count += int(xmlData[count_start + 11:count_end])
        price += int(xmlData[price_start + 12:price_end])
    return [count, price]

# 엑셀 파일 초기화
def setDataFile():
    workbook = Workbook()
    prevSheet = workbook.active
    workbook.remove(prevSheet)
    diagYms = getDiagYms()
    diagYms += diagYms
    diagYms = sorted(diagYms, reverse=True)
    sheet = workbook.create_sheet('datasheet')
    data = ['', '', '', '기간'] + diagYms
    sheet.append(data)
    data = ['약효분류군', '약효분류명', '시도', '시군구'] + (['사용량', '금액']*12)
    sheet.append(data)
    workbook.save('./datafile.xlsx')
    workbook.close()

# 엑셀 파일
def updateDataFile(sgguCode, sidoCode, meftDivNo, data):
    workbook = load_workbook('datafile.xlsx')
    sheet = workbook.active
    sidoNames = getSidoNames()
    sgguNames = getSgguNames()
    meftDiv = getMeftDiv()
    use = []
    for uses in data.values():
        use.append(uses[0])
        use.append(uses[1])
    sheet.append([meftDivNo, meftDiv[meftDivNo], sidoNames[sidoCode], sgguNames[sgguCode]]+use)
    print([meftDivNo, meftDiv[meftDivNo], sidoNames[sidoCode], sgguNames[sgguCode]]+use)
    workbook.save('./datafile.xlsx')
    workbook.close()

def updateDataSum(meftDivNo, sidoCode, count):
    workbook = load_workbook('datafile.xlsx', data_only=True)
    sheet = workbook['datasheet']
    last_row = sheet.max_row
    first_row = last_row - count + 2
    sums = np.zeros((1, 24))
    for row_data in sheet.iter_rows(min_row=first_row):
        if count <= 0:
            break
        sum = []
        iter = 0
        for cell in row_data:
            if iter >= 4:
                sum.append(int(cell.value))
            iter += 1
        count -= 1
        now_sum = np.array(sum).reshape(1, 24)
        sums += now_sum

    meftDiv = getMeftDiv()
    sidoName = getSidoNames()
    data = [meftDivNo, meftDiv[meftDivNo], sidoName[sidoCode], '전체']
    data += sums.tolist()[0]
    print(data)
    sheet.append(data)
    workbook.save('./datafile.xlsx')
    workbook.close()

# 데이터를 추출하여 저장하는 부분
# 현재는 텍스트파일로 저장하지만 추후에 excel파일로 저장하는 코드를 추가할 예정
def getData():
    sgguCodes = getSgguNames().keys()
    diagYms = getDiagYms()
    meftDivNos = getMeftDivNos()
    last_sidoCode = 0
    data = dict()
    iter = 0
    count = 0
    setDataFile()
    url = 'http://apis.data.go.kr/B551182/msupUserInfoService/getMeftDivAreaList'

    for meftDivNo in meftDivNos:
        iter += 1
        if iter > 3:
            return
        for sgguCode in sgguCodes:
            sidoCode = str(int(sgguCode)//10000 * 10000)
            for diagYm in diagYms:
                params = {
                    'serviceKey': '3n9Fletn7bWgtk9FliVuTIafybKUNzv79P5L/f53imL6Z+A/auwIqy26sfuxMl3aN9i//0J3ryLzOs/NDTiFog==',
                    'numOfRows': '10',
                    'pageNo': '1',
                    'diagYm': diagYm,
                    'meftDivNo': meftDivNo,
                    'insupTp': '0',
                    'cpmdPrscTp': '01',
                    'sidoCd': sidoCode,
                    'sgguCd': sgguCode}

                response = requests.get(url, params=params)
                response.encoding = 'utf-8'
                xmlData = response.text
                data[diagYm] = parse(xmlData)
            count += 1
            if last_sidoCode != 0 and last_sidoCode != sidoCode:
                updateDataSum(meftDivNo, last_sidoCode, count)
                count = 1
            updateDataFile(sgguCode, sidoCode, meftDivNo, data)
            last_sidoCode = sidoCode
            data.clear()

#main
print('============ START ============')
getData()
print('============= END =============')
